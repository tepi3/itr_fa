import io
import logging
from datetime import datetime, date, time
from core.utils import tax_round, parse_sort_date

logger = logging.getLogger(__name__)

# openpyxl is lazy-loaded to speed up app startup
openpyxl = None

def _get_openpyxl():
    global openpyxl
    if openpyxl is None:
        import openpyxl as oxl
        openpyxl = oxl
    return openpyxl

def parse_vested_datetime(date_val, time_val) -> tuple:
    """
    Parse Vested Excel date and time values into a datetime object for sorting 
    and a string formatted as dd/mm/yyyy for transactions.
    """
    if not date_val:
        return None, None

    # Parse date part
    if isinstance(date_val, (datetime, date)):
        date_str = date_val.strftime("%Y-%m-%d")
    else:
        date_str = str(date_val).strip().split(" ")[0]
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m/%d/%y", "%d/%m/%y", "%d-%b-%Y", "%d-%b-%y"):
            try:
                dt = datetime.strptime(date_str, fmt)
                date_str = dt.strftime("%Y-%m-%d")
                break
            except ValueError:
                pass
                
    # Parse time part
    if isinstance(time_val, time):
        time_str = time_val.strftime("%I:%M:%S %p")
    elif isinstance(time_val, datetime):
        time_str = time_val.strftime("%I:%M:%S %p")
    else:
        time_str = str(time_val).strip() if time_val is not None else "12:00:00 AM"

    # Combine date and time
    dt_str = f"{date_str} {time_str}"
    
    # Try parsing combined formats
    for fmt in ("%Y-%m-%d %I:%M:%S %p", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %I:%M %p", "%Y-%m-%d %H:%M"):
        try:
            dt_obj = datetime.strptime(dt_str, fmt)
            return dt_obj, dt_obj.strftime("%d/%m/%Y")
        except ValueError:
            pass

    # Fallback to date only
    try:
        dt_obj = datetime.strptime(date_str, "%Y-%m-%d")
        return dt_obj, dt_obj.strftime("%d/%m/%Y")
    except ValueError:
        return None, None

def _extract_trades_from_file(file_bytes: bytes, filename: str) -> list:
    """
    Extracts raw trade rows from a single Vested statements XLSX.
    Returns a list of dicts with raw trade data.
    """
    oxl = _get_openpyxl()
    try:
        wb = oxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Could not load Excel file {filename}: {str(e)}")

    if 'Trades' not in wb.sheetnames:
        raise ValueError(f"Sheet 'Trades' not found in file: {filename}")

    ws = wb['Trades']
    
    # Header detection
    headers = []
    header_row_idx = -1
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
        if "ticker" in row_lower and "activity" in row_lower and "quantity" in row_lower:
            headers = [str(c).strip() if c is not None else "" for c in row]
            header_row_idx = idx
            break

    if header_row_idx == -1:
        raise ValueError(f"Could not find valid column headers in 'Trades' sheet of {filename}")

    lower_headers = [h.lower() for h in headers]
    
    def find_idx(possible_names):
        for name in possible_names:
            if name in lower_headers:
                return lower_headers.index(name)
        return -1

    date_idx = find_idx(["date"])
    time_idx = find_idx(["time (in utc)", "time"])
    ticker_idx = find_idx(["ticker", "symbol"])
    activity_idx = find_idx(["activity", "type"])
    qty_idx = find_idx(["quantity", "qty"])
    price_idx = find_idx(["price per share (in usd)", "price per share", "price"])
    amount_idx = find_idx(["cash amount (in usd)", "cash amount", "amount"])
    comm_idx = find_idx(["commission charges (in usd)", "commission charges", "commission", "charges"])

    if date_idx == -1 or ticker_idx == -1 or activity_idx == -1 or qty_idx == -1:
        raise ValueError(f"Missing required columns in Vested Trades sheet in {filename}")

    trades = []
    
    # Read rows below header
    for row in list(ws.iter_rows(values_only=True))[header_row_idx + 1:]:
        min_needed = max(date_idx, ticker_idx, activity_idx, qty_idx)
        if len(row) <= min_needed:
            continue

        ticker = str(row[ticker_idx] or "").strip().upper()
        if not ticker:
            continue

        date_val = row[date_idx]
        time_val = row[time_idx] if time_idx != -1 else None
        
        dt_obj, date_str = parse_vested_datetime(date_val, time_val)
        if not dt_obj:
            continue

        activity = str(row[activity_idx] or "").strip().lower()
        if activity not in ("buy", "sell"):
            continue

        try:
            qty = abs(float(str(row[qty_idx]).replace(",", "")))
            if qty == 0:
                continue
        except (ValueError, TypeError):
            continue

        # Extract price / amount info
        price_raw = None
        if price_idx != -1 and price_idx < len(row):
            try:
                price_raw = float(str(row[price_idx]).replace("$", "").replace(",", ""))
            except (ValueError, TypeError):
                pass

        cash_amount = None
        if amount_idx != -1 and amount_idx < len(row):
            try:
                cash_amount = float(str(row[amount_idx]).replace("$", "").replace(",", ""))
            except (ValueError, TypeError):
                pass

        comm = 0.0
        if comm_idx != -1 and comm_idx < len(row):
            try:
                comm = float(str(row[comm_idx]).replace("$", "").replace(",", ""))
            except (ValueError, TypeError):
                pass

        # Compute price: cash amount / quantity includes commissions.
        # Fallback to execution price if cash amount is missing.
        if cash_amount is not None and qty > 0:
            price = tax_round(cash_amount / qty, 2)
        elif price_raw is not None:
            price = tax_round(price_raw, 2)
        else:
            continue

        trades.append({
            "ticker": ticker,
            "dt_obj": dt_obj,
            "date": date_str,
            "activity": activity,
            "qty": qty,
            "price": price,
            "commission": comm
        })

    return trades

def process_vested_files(file_list: list, calendar_year: int) -> dict:
    """
    Parses multiple Vested Transaction statement XLSX files, 
    and builds FIFO lot-matched transactions for a target calendar year.

    Args:
        file_list: List of (file_bytes, filename) tuples.
        calendar_year: The target calendar year for the portfolio.

    Returns:
        dict: {
            "transactions": List of flat linked BUY/SELL transactions for target CY.
            "unmatched_sells": List of sells that couldn't be matched to any buy lot.
            "skipped_count": Number of trades skipped (after calendar year).
        }
    """
    all_trades = []
    for file_bytes, filename in file_list:
        file_trades = _extract_trades_from_file(file_bytes, filename)
        all_trades.extend(file_trades)

    # Sort chronologically by datetime object
    all_trades.sort(key=lambda t: t["dt_obj"])

    lots_by_symbol = {}
    unmatched_sells = []
    skipped_count = 0

    for trade in all_trades:
        trade_year = trade["dt_obj"].year
        if trade_year > calendar_year:
            skipped_count += 1
            continue

        sym = trade["ticker"]
        qty = tax_round(trade["qty"], 6)
        price = trade["price"]
        date_str = trade["date"]

        if trade["activity"] == "buy":
            if sym not in lots_by_symbol:
                lots_by_symbol[sym] = []
            lots_by_symbol[sym].append({
                "buy_date": date_str,
                "buy_price": price,
                "quantity": qty,
                "sells": []
            })
        elif trade["activity"] == "sell":
            symbol_lots = lots_by_symbol.get(sym, [])
            sell_qty_left = qty

            for lot in symbol_lots:
                sold_qty = sum(s["quantity"] for s in lot["sells"])
                available = tax_round(lot["quantity"] - sold_qty, 6)
                if available <= 0:
                    continue

                matched = tax_round(min(sell_qty_left, available), 6)
                lot["sells"].append({
                    "sell_date": date_str,
                    "quantity": matched,
                    "sell_price": price
                })
                sell_qty_left = tax_round(sell_qty_left - matched, 6)
                if sell_qty_left <= 0:
                    break

            if sell_qty_left > 0:
                unmatched_sells.append({
                    "symbol": sym,
                    "date": date_str,
                    "qty": sell_qty_left,
                    "sell_price": price
                })
                logger.warning(
                    f"Vested lot match failed for {sym} sell on {date_str}: "
                    f"qty={sell_qty_left}, sell_price={price}"
                )

    # Emit linked transactions for the calendar year
    transactions = []
    for sym, lots in lots_by_symbol.items():
        for lot in lots:
            buy_date = lot["buy_date"]
            buy_price = lot["buy_price"]
            original_qty = lot["quantity"]

            # Calculate sold quantity before target year
            pre_cy_sold = 0.0
            cy_sells = []
            for sell in lot["sells"]:
                try:
                    sell_d = parse_sort_date(sell["sell_date"])
                    if sell_d.year < calendar_year:
                        pre_cy_sold += sell["quantity"]
                    elif sell_d.year == calendar_year:
                        cy_sells.append(sell)
                except Exception:
                    cy_sells.append(sell)

            # Remaining quantity carried forward into target CY
            carried_qty = tax_round(original_qty - pre_cy_sold, 6)
            if carried_qty <= 0:
                continue  # Lot fully sold before this CY

            # Emit BUY transaction for carried-forward quantity
            transactions.append({
                "type": "BUY",
                "date": buy_date,
                "symbol": sym,
                "qty": carried_qty,
                "price": buy_price
            })

            # Emit linked SELL transactions for sells during target CY
            for sell in cy_sells:
                transactions.append({
                    "type": "SELL",
                    "date": sell["sell_date"],
                    "symbol": sym,
                    "qty": sell["quantity"],
                    "price": sell["sell_price"],
                    "buy_date": buy_date,
                    "buy_price": buy_price
                })

    # Sort transactions chronologically
    transactions.sort(key=lambda t: parse_sort_date(t["date"]))

    logger.info(
        f"Vested extraction: {len(transactions)} transactions, "
        f"{len(unmatched_sells)} unmatched sells, "
        f"{skipped_count} skipped (after CY)"
    )

    return {
        "transactions": transactions,
        "unmatched_sells": unmatched_sells,
        "skipped_count": skipped_count
    }
