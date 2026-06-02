import io
import logging
import uuid
from datetime import datetime
from core.utils import tax_round, parse_sort_date

# openpyxl is lazy-loaded to speed up app startup
openpyxl = None

def _get_openpyxl():
    global openpyxl
    if openpyxl is None:
        import openpyxl as oxl
        openpyxl = oxl
    return openpyxl

logger = logging.getLogger(__name__)


def parse_date(date_val) -> str:
    """Parse common date formats into dd/mm/yyyy."""
    if not date_val or str(date_val).strip() in ("--", "NA", "N/A", ""):
        return None
    if isinstance(date_val, datetime):
        return date_val.strftime("%d/%m/%Y")
    date_str = str(date_val).strip().split(" ")[0]
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%d-%b-%y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(date_str, fmt).strftime("%d/%m/%Y")
        except ValueError:
            pass
    return None


def _get_year(date_str: str) -> int:
    """Extract year from a dd/mm/yyyy date string."""
    if not date_str:
        return 0
    try:
        parts = date_str.split("/")
        y = int(parts[-1])
        if y < 100:
            y += 2000
        return y
    except (ValueError, IndexError):
        return 0


def _clean_float(val) -> float:
    """Clean currency/number values from Excel cells."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return tax_round(float(val), 6)
    try:
        clean_str = str(val).replace("$", "").replace(",", "").strip()
        return tax_round(float(clean_str), 6)
    except (ValueError, TypeError):
        return 0.0


def _extract_company_name(ws) -> str:
    """Extract company name from the title row (row 1) of a sheet.
    
    Title format: 'Company Name, Inc - RSU Sales (report run on ...)'
    """
    try:
        title = ws.cell(row=1, column=1).value
        if title and isinstance(title, str) and " - " in title:
            return title.split(" - ")[0].strip()
    except Exception:
        pass
    return ""


def _is_data_row(row) -> bool:
    """Check if a row is an actual data row (not a header, footer, or disclaimer)."""
    first_cell = row[0]
    if first_cell is None:
        return False
    val = str(first_cell).strip().lower()
    # Skip disclaimer/footer rows
    if val.startswith("all currencies") or val.startswith("***"):
        return False
    # Data rows have a numeric employee number in the first column
    try:
        int(str(first_cell).strip())
        return True
    except (ValueError, TypeError):
        return False


def _parse_releases_report(ws, target_year: int, ticker: str):
    """Parse the Releases Report sheet for RSU vesting events → BUY transactions.
    
    Header layout (row 2):
      [0] Employee Number, ..., [6] Release Date, [7] Release Price, ...,
      [9] Issued (Units), [10] Released, [11] Net Withheld, [12] Sold, ...
    
    Row 3 is a sub-header row for the Units columns — skip it.
    Data starts at row 4 (0-indexed row 3).
    """
    transactions = []
    skipped = 0

    # Find header row and column indices
    headers = [cell.value for cell in ws[2]]  # Row 2 is the header
    
    # Map column indices by header name
    col_map = {}
    for idx, h in enumerate(headers):
        if h is not None:
            col_map[str(h).strip().lower()] = idx
    
    release_date_idx = col_map.get("release date")
    release_price_idx = col_map.get("release price")
    
    # Units sub-columns: row 2 has "Units" spanning, row 3 has sub-headers
    # But the sub-headers in row 3 are: Issued, Released, Net Withheld, Sold
    # We need to find "Issued" from row 3
    sub_headers = [cell.value for cell in ws[3]]  # Row 3 is the sub-header
    issued_idx = None
    for idx, h in enumerate(sub_headers):
        if h is not None and str(h).strip().lower() == "issued":
            issued_idx = idx
            break
    
    if release_date_idx is None or release_price_idx is None or issued_idx is None:
        logger.warning("Releases Report: Could not find required columns. "
                       f"release_date={release_date_idx}, release_price={release_price_idx}, issued={issued_idx}")
        return transactions, skipped

    # Data starts at row 4 (ws rows are 1-indexed, so row index 4)
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not _is_data_row(row):
            continue

        release_date = parse_date(row[release_date_idx])
        if not release_date:
            continue

        buy_year = _get_year(release_date)
        if buy_year > target_year:
            skipped += 1
            continue

        qty = _clean_float(row[issued_idx])
        price = _clean_float(row[release_price_idx])

        if qty <= 0 or price <= 0:
            continue

        transactions.append({
            "type": "BUY",
            "date": release_date,
            "symbol": ticker,
            "qty": tax_round(qty, 6),
            "price": tax_round(price, 2),
        })

    return transactions, skipped


def _parse_espp_purchase(ws, target_year: int, ticker: str):
    """Parse the ESPP Purchase sheet → BUY transactions.
    
    Header layout (row 2):
      [0] Employee Number, ..., [5] Purchase Date, [6] Purchase Date FMV,
      [7] Purchase Price, [8] Shares Purchased, [9] Total Purchase Price
    """
    transactions = []
    skipped = 0

    headers = [cell.value for cell in ws[2]]
    col_map = {}
    for idx, h in enumerate(headers):
        if h is not None:
            col_map[str(h).strip().lower()] = idx

    purchase_date_idx = col_map.get("purchase date")
    fmv_idx = col_map.get("purchase date fmv")
    shares_idx = col_map.get("shares purchased")

    if purchase_date_idx is None or fmv_idx is None or shares_idx is None:
        logger.warning("ESPP Purchase: Could not find required columns. "
                       f"purchase_date={purchase_date_idx}, fmv={fmv_idx}, shares={shares_idx}")
        return transactions, skipped

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not _is_data_row(row):
            continue

        buy_date = parse_date(row[purchase_date_idx])
        if not buy_date:
            continue

        buy_year = _get_year(buy_date)
        if buy_year > target_year:
            skipped += 1
            continue

        qty = _clean_float(row[shares_idx])
        price = _clean_float(row[fmv_idx])

        if qty <= 0 or price <= 0:
            continue

        transactions.append({
            "type": "BUY",
            "date": buy_date,
            "symbol": ticker,
            "qty": tax_round(qty, 6),
            "price": tax_round(price, 2),
        })

    return transactions, skipped


def _parse_rsu_sales(ws, target_year: int, ticker: str):
    """Parse the RSU Sales sheet → linked SELL transactions.
    
    Header layout (row 2):
      [0] Employee Number, ..., [3] Sale Date, [4] Acquisition Date,
      [5] Sale Price, [6] Sale Net Proceeds, [7] Sale Quantity,
      [8] Cost Basis per Share, [9] Cost Basis For Lot, [10] Gain From Sale
    """
    transactions = []
    skipped = 0

    headers = [cell.value for cell in ws[2]]
    col_map = {}
    for idx, h in enumerate(headers):
        if h is not None:
            col_map[str(h).strip().lower()] = idx

    sale_date_idx = col_map.get("sale date")
    acq_date_idx = col_map.get("acquisition date")
    sale_price_idx = col_map.get("sale price")
    sale_net_proceeds_idx = col_map.get("sale net proceeds")
    sale_qty_idx = col_map.get("sale quantity")
    cost_basis_idx = col_map.get("cost basis per share")

    if sale_date_idx is None or acq_date_idx is None or sale_price_idx is None or sale_qty_idx is None:
        logger.warning("RSU Sales: Could not find required columns.")
        return transactions, skipped

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not _is_data_row(row):
            continue

        sell_date = parse_date(row[sale_date_idx])
        buy_date = parse_date(row[acq_date_idx])
        if not sell_date or not buy_date:
            continue

        sell_year = _get_year(sell_date)
        buy_year = _get_year(buy_date)

        # Skip shares acquired AFTER the target year
        if buy_year > target_year:
            skipped += 1
            continue
        # Skip shares sold BEFORE the target year
        if sell_year < target_year:
            skipped += 1
            continue

        qty = _clean_float(row[sale_qty_idx])
        
        # Use Net Proceeds if available, otherwise fallback to Sale Price
        if sale_net_proceeds_idx is not None:
            net_proceeds = _clean_float(row[sale_net_proceeds_idx])
            sell_price = tax_round(net_proceeds / qty, 2)
        else:
            sell_price = _clean_float(row[sale_price_idx])
            
        buy_price = _clean_float(row[cost_basis_idx]) if cost_basis_idx is not None else 0.0

        if qty <= 0:
            continue

        transactions.append({
            "type": "SELL",
            "date": sell_date,
            "symbol": ticker,
            "qty": tax_round(qty, 6),
            "price": tax_round(sell_price, 2),
            "buy_date": buy_date,
            "buy_price": tax_round(buy_price, 2),
        })

    return transactions, skipped


def _parse_espp_sales(ws, target_year: int, ticker: str):
    """Parse the ESPP Sales sheet → linked SELL transactions.
    
    Header layout (row 2):
      [0] Employee Number, ..., [3] Sale Date, [4] Subscription Date,
      [5] Subscription Date FMV, [6] Purchase Date, [7] Purchase Date FMV,
      [8] Purchase Price, [9] Sale Quantity, [10] Sale Price,
      [11] Sale Net Proceeds, [12] Gain From Sale
    """
    transactions = []
    skipped = 0

    headers = [cell.value for cell in ws[2]]
    col_map = {}
    for idx, h in enumerate(headers):
        if h is not None:
            col_map[str(h).strip().lower()] = idx

    sale_date_idx = col_map.get("sale date")
    purchase_date_idx = col_map.get("purchase date")
    purchase_fmv_idx = col_map.get("purchase date fmv")
    sale_qty_idx = col_map.get("sale quantity")
    sale_price_idx = col_map.get("sale price")

    if sale_date_idx is None or purchase_date_idx is None or sale_qty_idx is None or sale_price_idx is None:
        logger.warning("ESPP Sales: Could not find required columns.")
        return transactions, skipped

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not _is_data_row(row):
            continue

        sell_date = parse_date(row[sale_date_idx])
        buy_date = parse_date(row[purchase_date_idx])
        if not sell_date or not buy_date:
            continue

        sell_year = _get_year(sell_date)
        buy_year = _get_year(buy_date)

        if buy_year > target_year:
            skipped += 1
            continue
        if sell_year < target_year:
            skipped += 1
            continue

        qty = _clean_float(row[sale_qty_idx])
        sell_price = _clean_float(row[sale_price_idx])
        # Use Purchase Date FMV as cost basis to match the acquisition lot
        buy_price = _clean_float(row[purchase_fmv_idx]) if purchase_fmv_idx is not None else 0.0

        if qty <= 0:
            continue

        transactions.append({
            "type": "SELL",
            "date": sell_date,
            "symbol": ticker,
            "qty": tax_round(qty, 6),
            "price": tax_round(sell_price, 2),
            "buy_date": buy_date,
            "buy_price": tax_round(buy_price, 2),
        })

    return transactions, skipped


# --- Entry Point ---
def process_morgan_stanley_file(file_bytes: bytes, filename: str,
                        target_year: int = 2025,
                        ticker_symbol: str = "",
                        company_info_map: dict = None) -> dict:
    """
    Parse a Morgan Stanley Share Sale Cost Basis Report XLSX and extract transactions.
    
    The file has four sheets:
      - Releases Report  → RSU vesting events (BUY)
      - ESPP Purchase    → ESPP purchase events (BUY)
      - RSU Sales        → RSU sale events (linked SELL)
      - ESPP Sales       → ESPP sale events (linked SELL)
    
    Returns: {transactions: [...], skipped_count: int, company_name: str}
    """
    oxl = _get_openpyxl()
    wb = oxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    ticker = ticker_symbol.strip().upper() if ticker_symbol else ""
    company_name = ""

    all_transactions = []
    total_skipped = 0

    # Sheet name mapping (case-insensitive lookup)
    sheet_map = {name.lower(): name for name in wb.sheetnames}

    # 1. Parse Releases Report (RSU vesting → BUY)
    releases_key = None
    for key in sheet_map:
        if "release" in key:
            releases_key = key
            break
    if releases_key:
        ws = wb[sheet_map[releases_key]]
        if not company_name:
            company_name = _extract_company_name(ws)
        if not ticker:
            # Fallback: use company name as ticker (user should provide real one)
            ticker = company_name.split(",")[0].strip().upper()[:10] if company_name else "UNKNOWN"
        txs, skipped = _parse_releases_report(ws, target_year, ticker)
        all_transactions.extend(txs)
        total_skipped += skipped

    # 2. Parse ESPP Purchase (→ BUY)
    espp_purchase_key = None
    for key in sheet_map:
        if "espp" in key and "purchase" in key:
            espp_purchase_key = key
            break
    if espp_purchase_key:
        ws = wb[sheet_map[espp_purchase_key]]
        if not company_name:
            company_name = _extract_company_name(ws)
        txs, skipped = _parse_espp_purchase(ws, target_year, ticker)
        all_transactions.extend(txs)
        total_skipped += skipped

    # 3. Parse RSU Sales (→ linked SELL)
    rsu_sales_key = None
    for key in sheet_map:
        if "rsu" in key and "sale" in key:
            rsu_sales_key = key
            break
    if rsu_sales_key:
        ws = wb[sheet_map[rsu_sales_key]]
        txs, skipped = _parse_rsu_sales(ws, target_year, ticker)
        all_transactions.extend(txs)
        total_skipped += skipped

    # 4. Parse ESPP Sales (→ linked SELL)
    espp_sales_key = None
    for key in sheet_map:
        if "espp" in key and "sale" in key:
            espp_sales_key = key
            break
    if espp_sales_key:
        ws = wb[sheet_map[espp_sales_key]]
        txs, skipped = _parse_espp_sales(ws, target_year, ticker)
        all_transactions.extend(txs)
        total_skipped += skipped

    # Sort all transactions chronologically
    all_transactions.sort(key=lambda tx: parse_sort_date(tx["date"]))

    logger.info(f"Morgan Stanley extraction: {len(all_transactions)} transactions found, "
                f"{total_skipped} skipped, ticker={ticker}")

    return {
        "transactions": all_transactions,
        "skipped_count": total_skipped,
        "company_name": company_name,
    }
